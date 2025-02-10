import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import re
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings("ignore", category=FutureWarning)


def select_directory():
    """Open a dialog to select a directory and return the selected path."""
    root = tk.Tk()
    root.withdraw()
    return filedialog.askdirectory()


def get_text_files(directory):
    """List all text files in the given directory.

    Args:
        directory (str): Path to the directory.

    Returns:
        list: A list of filenames (str) that end with '.txt'.
    """
    return [file for file in os.listdir(directory) if file.endswith('.txt')]


def custom_sort_key(filename):
    """Generate a sort key for each filename based on predefined logic.

    Args:
        filename (str): The filename to generate a sort key for.

    Returns:
        tuple: A tuple containing the primary and secondary sort keys.
    """
    parts = filename.split(' ')
    primary_sort = parts[0]
    secondary_sort_value = 0 if parts[1] == "Clicks-analyzed.txt" else float(
        parts[1].split('.')[0])
    return primary_sort, secondary_sort_value


def process_file(directory, text_file):
    """Read and process a text file, converting it to a DataFrame.

    Args:
        directory (str): The directory containing the file.
        text_file (str): The filename to process.

    Returns:
        DataFrame: A pandas DataFrame containing the processed file data.
    """
    data = []
    with open(os.path.join(directory, text_file), 'r') as file:
        for line in file:
            processed_line = [item for sublist in [elem.split(',') for elem in
                                                   line.strip().split('\t')] for
                              item in sublist]
            data.append(processed_line)
    return pd.DataFrame(data)


def extract_info(df, text_file):
    """Extract metadata from the dataframe and filename.

    Args:
        df (DataFrame): The pandas DataFrame containing file data.
        text_file (str): The filename, used to extract additional info.

    Returns: tuple: Contains extracted information including the filename ID,
    threshold, frequency, and threshold method.
    """

    threshold = df.iloc[0, 0].split(':')[-1].strip()
    frequency_raw = df.iloc[1, 0].split(':')[-1].strip()

    frequency = frequency_raw.split(' ')[1] if ' ' in frequency_raw else (
        frequency_raw)
    frequency = 'Clicks' if frequency == '0.00' else \
        f"{float(frequency):.2g} kHz"
    if 'Noise' in df.iloc[2,0]:
        threshold_method_raw = df.iloc[3, 0].split(':')[-1].strip()
    else:
        threshold_method_raw = df.iloc[2, 0].split(':')[-1].strip()

    if threshold_method_raw == '':
        threshold_method = 'None'
    elif threshold_method_raw == 'manual':
        threshold_method = 'Manual'
    else:
        threshold_method = threshold_method_raw.capitalize() + ')'

    base_name, file_extension = os.path.splitext(text_file)
    base_name_parts = base_name.split('-')
    isolated_filename = ' '.join(base_name_parts[:-1])
    isolated_filename = isolated_filename.split(' ')
    isolated_filename = ' '.join(isolated_filename[:-2]) \
        if ('kHz' in isolated_filename) else ' '.join(isolated_filename[:-1])
    isolated_filename = isolated_filename.replace(" ", "-")
    return isolated_filename, threshold, frequency, threshold_method


def process_dataframe(df, name, threshold, frequency, threshold_method):
    """Process and clean the DataFrame by setting up correct columns and
    metadata.

    Args:
        df (DataFrame): The DataFrame to process.
        name (str): The ID derived from the filename.
        threshold (str): The extracted threshold value.
        frequency (str): The frequency value.
        threshold_method (str): The method used for threshold determination.

    Returns:
        DataFrame: The processed DataFrame with additional columns and metadata.
    """
    if 'Noise' in df.iloc[2,0]:
        df = df.drop(range(7)).reset_index(drop=True)
    else:
        df = df.drop(range(6)).reset_index(drop=True)

    df.columns = df.iloc[0]
    df = df.drop(df.index[0])
    df['ID'] = name
    df['Threshold'] = threshold
    df['Frequency'] = frequency
    df['Threshold method'] = threshold_method
    df = df.applymap(lambda x: pd.to_numeric(x, errors='ignore'))
    df['CorrCoef'] = df['CorrCoef'].fillna(1)
    return df


def calculate_amplitudes(df):
    """Calculate wave amplitudes and update the DataFrame accordingly.

    Args:
        df (DataFrame): The DataFrame to calculate wave amplitudes for.

    Returns:
        DataFrame: Updated DataFrame with wave amplitude calculations.
    """
    p_pattern = r'P(\d+) Amplitude'
    n_pattern = r'N(\d+) Amplitude'
    for column in df.columns:
        p_match = re.match(p_pattern, column)
        if p_match:
            wave_number = p_match.group(1)
            n_column = f'N{wave_number} Amplitude'
            if n_column in df.columns:
                df[f'Wave {wave_number} Amplitude'] = df[column] - df[
                    n_column]
                df[f'Wave {wave_number} Amplitude'] = df[
                    f'Wave {wave_number} Amplitude'].clip(0)
    return df


def reorder_columns(df):
    """Reorder DataFrame columns based on specific criteria.

    Args:
        df (DataFrame): The DataFrame whose columns need to be reordered.

    Returns:
        DataFrame: The DataFrame with reordered columns.
    """
    new_order = []

    for col in df.columns:
        if "Wave" not in col:
            new_order.append(col)
            if "N" in col and "Latency" not in col and "Amplitude" in col:
                wave_num = col.split(" ")[0][1:]
                wave_col = f'Wave {wave_num} Amplitude'
                if wave_col in df.columns:
                    new_order.append(wave_col)

    df = df[new_order]

    new_order2 = ['ID', 'Frequency', 'Threshold', 'Threshold method',
                  'Level', 'CorrCoef'] + [col for col in df.columns if
                                          col not in ['ID', 'Threshold',
                                                      'Frequency',
                                                      'Threshold method',
                                                      'Level',
                                                      'CorrCoef']]
    df = df[new_order2]
    df = df.rename(
        columns={'Level': 'Intensity', 'method': 'Method'}).sort_values(
        by='Intensity', ascending=True)
    return df


def get_numeric_columns(df):
    """Identify and return numeric columns, excluding specific columns.

    Args:
        df (DataFrame): The DataFrame to examine.

    Returns: list: A list of column names that are numeric, excluding
    specified columns.
    """
    exclude_columns = ['Threshold', 'Intensity', 'CorrCoef', '0.3msec Avg',
                       '0.3msec StDev']
    return [column for column in df.columns if
            column not in exclude_columns and pd.api.types.is_numeric_dtype(
                df[column])]


def adjust_values(row, numeric_columns):
    """Adjust row values based on threshold conditions.

    Args:
        row (Series): A pandas Series representing a row in the DataFrame.
        numeric_columns (list): List of numeric column names to adjust.

    Returns:
        Series: The adjusted pandas Series (row).
    """
    threshold = float(row['Threshold']) if row['Threshold'] != 'None' else 0.0
    if float(row['Intensity']) < threshold:
        for col in numeric_columns:
            row[col] = 0
    return row


def export_to_excel(master_df, directory):
    """Export the master DataFrame to an Excel file in the selected directory.

    Args:
        master_df (DataFrame): The master DataFrame to export.
        directory (str): The directory to save the Excel file in.
    """
    full_path = os.path.join(directory, "Merged_dataframe.xlsx")
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        master_df.to_excel(writer, sheet_name='Sheet1', index=False)
        format_excel_file(writer, master_df)


def format_excel_file(writer, master_df):
    """Format the Excel file, adjusting columns, and applying styles.

    Args: writer (ExcelWriter): The writer instance to apply formatting with.
    master_df (DataFrame): The DataFrame being written to Excel, for reference.
    """
    worksheet = writer.sheets['Sheet1']
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) + 2 for cell in column_cells)
        worksheet.column_dimensions[
            get_column_letter(column_cells[0].column)].width = length
    table_range = (f"A1:{get_column_letter(worksheet.max_column)}"
                   f"{worksheet.max_row}")
    table = Table(displayName="DataTable", ref=table_range,
                  tableStyleInfo=TableStyleInfo(name="TableStyleLight1",
                                                showFirstColumn=False,
                                                showLastColumn=False,
                                                showRowStripes=True,
                                                showColumnStripes=False))
    worksheet.add_table(table)
    for col_idx, column in enumerate(master_df.columns, start=1):
        if pd.api.types.is_numeric_dtype(master_df[column]):
            cell_range = (f"{get_column_letter(col_idx)}2:"
                          f"{get_column_letter(col_idx)}{worksheet.max_row}")
            rule = ColorScaleRule(start_type='min', start_color='F0F0F0',
                                  end_type='max', end_color='A0A0A0',
                                  mid_type='percentile', mid_value=50,
                                  mid_color='C8C8C8')
            worksheet.conditional_formatting.add(cell_range, rule)


def main():
    """Main function to execute the program workflow."""
    directory = select_directory()
    text_files = sorted(get_text_files(directory), key=custom_sort_key)
    master_df = pd.DataFrame()
    for text_file in text_files:
        df = process_file(directory, text_file)
        name, threshold, frequency, threshold_method = extract_info(df,
                                                                    text_file)
        df = process_dataframe(df, name, threshold, frequency, threshold_method)
        df = calculate_amplitudes(df)
        df = reorder_columns(df)
        numeric_columns = get_numeric_columns(df)
        df = df.apply(lambda row: adjust_values(row, numeric_columns), axis=1)
        master_df = pd.concat([master_df, df], ignore_index=True)
        print("Transforming " + name + " " + frequency + "...")
    export_to_excel(master_df, directory)
    print("Processing complete. The Excel file has been saved.")


if __name__ == "__main__":
    main()
